import streamlit as st
import time
from datetime import datetime,timedelta
import json
import urllib.request
import requests
#from pyzbar.pyzbar import decode
import qrcode
import openpyxl

#path = "C:\\Users\\tapan\\OneDrive\\Desktop\\qrbot\\"


import cv2

st.title('Welcome to QR code coder decoder')
polltime=2
pollgap=2
u2=""
replyto="5560841599"
looper=0
message = ""
updatetext1=0
token="6158093548:AAEw6MXroTQMxxPGc6gqvgzSh5D0j3Jbm3Q"
headers = {"accept": "application/json","content-type": "application/json"}
photourl="https://api.telegram.org/bot"+token+"/sendPhoto"
texturl="https://api.telegram.org/bot"+token+"/sendMessage"
meurl ="https://api.telegram.org/bot"+token+"/getMe"
pollurl="https://api.telegram.org/bot"+token+"/getUpdates"
filedownload="https://api.telegram.org/file/bot"+token+"/"
fileurl="https://api.telegram.org/bot"+token+"/getfile"
healthmessage="This is health message of QR coder/Decoder BOT. It is running fine and you are receiving this message every 5 minutes "
healthtime = int(time.time())
advertise1 = "Get Your Free Personal Assistant- Eazyai is all in one AI powered Personal Assistant which can create SEO-optimized and unique content for your blogs, ads, emails and website 10X faster, generate Images, convert Text to Speech and Speech to Text and enable you to interact with customized AI Chat Bots to save your precious time and resources. https://eazyai.io"

#advertise1 = "Visit https://tapanaitools.com to find free AI enabled seo tools to make your life easy "
#advertise1 = "Earn daily income through revolutionary Telegram AI bot that generates up to 10% with UTBAI token:"+"\nhttps://t.me/unitedtelebot?start=5779577400"
advertise=advertise1
#advertise= "Visit https://tapanaitools.com to find more such tools to make your life easy "

user = "user1.xlsx"
wb_obj = openpyxl.load_workbook(user)
sheet_obj = wb_obj.active
while looper==0:
    cn = 0
    mime=""
    messagelink = ""
    updatefile = open("updateid.txt", 'r+')
    updatetext = updatefile.read()
    updatefile.close()
    #Starting Long Polling
    lastupdate=int(updatetext)
    payloadpoll = {"offset":lastupdate,"limit": 50,"timeout": pollgap}
    response = requests.post(pollurl, json=payloadpoll, headers=headers)
    f = urllib.request.urlopen(pollurl)
    data=json.load(f)
    b=data

    c=str(b)

    lenc=len(c)
    print(c)
    print(lenc)
    if lenc<50:
        print("No New message "+str(datetime.now().strftime("%H:%M:%S")))
    if lenc>50:
      print("Message received at: "+str(datetime.now().strftime("%H:%M:%S")))
      tempstr = c[0:lenc]
      latid=""
      longid=""
      while cn==0:
       downloadlink=""
       usernametext=""
       usernamepos1=tempstr.find("username")
       if usernamepos1!=-1:
         usernamepos2=tempstr.find("language_code")
         usernametext=tempstr[usernamepos1+12:usernamepos2-4]
       filepos1 = tempstr.rfind("file_id':")
       if filepos1 != -1:

           filepos2 = tempstr.rfind("'file_unique_id':")
           fileid = tempstr[filepos1 + 11:filepos2 - 3]
           print(fileid)
           payload = {"file_id": fileid}
           response = requests.post(fileurl, json=payload, headers=headers)
           ok=response.text
           print(response.text)

           if  ok.find("true")!=-1:
               linkpos1 = ok.find("file_path")
               linkpos2 = ok.find("}}")
               link=ok[linkpos1 + 12:linkpos2 - 1]
               print(link)
               finddot= link.rfind(".")
               if finddot!=-1:
                   mime=link[finddot:len(link)]
                   print("mime:"+mime)
                   if mime != ".jpg" and mime != ".png" and mime != ".jpeg" and mime != ".bmp" and mime != ".webp" and mime =="":
                       messagelink="This is not a valid Image file. Please check and send Image file only"
               downloadlink=filedownload+link
               print(downloadlink)

       messagepostemp=0
       updatepos1 = tempstr.find("update_id':")
       print("updatepos1",updatepos1)
       if updatepos1==-1:
        break
       namepos = tempstr.find("is_bot': False, 'first_name':")
       namepos1 = tempstr.find("last_name")
       namepos2 = tempstr.find("username")
       namepos3 = tempstr.find("language_code")
       print("namepos", namepos)
       print("namepos1", namepos1)
       numberpos = tempstr.find("from")
       print("numberpos",numberpos)
       datepos = tempstr.find("date':")
       print("datepos",datepos)
       endpos1 = tempstr.find("}}, {")
       print("endpos1",endpos1)
       endpos2 = tempstr.find("}}]}")
       messagepos = tempstr.find("text")
       if messagepos-datepos>20:
         messagepostemp =messagepos
         messagepos=datepos+20
       print("messagepos",messagepos)
       updatetext1 = tempstr[updatepos1 + 12:updatepos1 + 21]
       print(updatetext1)
       numbertext = tempstr[numberpos + 14:numberpos + 24]
       print(numbertext)
       print(namepos1)

       if namepos1 == -1 :
           nametext = tempstr[namepos + 31:namepos3 - 4]

       if namepos1 != -1:
           nametext = tempstr[namepos + 31:namepos1 - 4]

       if namepos1 == -1 and namepos2!=-1 :
           nametext = tempstr[namepos + 31:namepos2 - 4]

       if namepos1 == -1 and namepos2 == -1 and namepos3!=-1:
           nametext = tempstr[namepos + 31:namepos3 - 4]

       print("Name: " + nametext)
       datetext = tempstr[datepos + 7:datepos + 17]
       dateconfirm = datetext[2:10]
       datetext=(datetime.fromtimestamp(int(datetext)).strftime('%Y-%m-%d %H:%M:%S'))
       print(datetext)
       if endpos1!=-1:
        messagetext=tempstr[messagepos+8:endpos1-1]
        if messagepostemp-datepos>20:
          messagetext="No Text"
       if endpos1==-1:
        messagetext=tempstr[messagepos+8:endpos2-1]
        if messagepostemp-datepos>20:
          messagetext="No Text"
       #print(messagetext)
       messagetext = messagetext.replace("\\xa0", " ")

       print(messagetext)
       print("----------------")

       print("username: "+usernametext)

       apos=nametext.find(",")
       if apos!=-1:
           u1=nametext[0:apos-1]
       else:
        u1 = nametext

       if u1 != u2:
           alert = u1 + " has just logged in\n\n"
           alert1 = "<a href='tg://user?id=" + numbertext + "'>Click to chat here</a>"
           if usernametext!="":
               alert1=alert1+"\n\nor "+"@"+usernametext
           payloadtext = {"text": alert + alert1, "parse_mode": "html", "disable_web_page_preview": False,
                          "disable_notification": False, "reply_to_message_id": None, "chat_id": replyto}
           response = requests.post(texturl, json=payloadtext, headers=headers)
           u2 = u1
       callbackpos=messagetext.find("'data':")
       callbacktext=messagetext[callbackpos+9:len(messagetext)]
       print("callbacktext:",callbacktext)
       numbertext=numbertext.rstrip(",")
       ab1 = 0
       roww1 = 2
       while ab1 == 0:

           cell_obj1 = sheet_obj.cell(row=roww1, column=1)
           cell_obj2 = sheet_obj.cell(row=roww1, column=2)
           cell_obj3 = sheet_obj.cell(row=roww1, column=3)

           if str(numbertext)==str(cell_obj1.value):
              break
           elif numbertext!=cell_obj1.value and  (cell_obj1.value==None or cell_obj1.value==""):
             cell_obj1.value = numbertext
             cell_obj2.value = datetext
             cell_obj3.value = u1
             wb_obj.save(user)
             break
           roww1=roww1+1
       # userfile = open(path + "user.txt", 'r',encoding='utf-8', errors='ignore')
       # usertext=userfile.read()
       # userfile.close
       # find1=usertext.find(numbertext)
       # if find1==-1:
       #  userfile = open(path + "user.txt", 'a' ,encoding='utf-8', errors='ignore')
       #  userfile.write(numbertext+","+datetext+","+u1+"\n")
       #  userfile.close()

       if messagelink!="":
           payloadtext = {"text": messagelink, "parse_mode": "html",
                          "disable_web_page_preview": False,
                          "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
           response = requests.post(texturl, json=payloadtext, headers=headers)
           break
       if downloadlink != "":
           file_url = downloadlink
           r = requests.get(file_url, stream=True)
           with open("decoded"+mime, 'wb') as f:
               f.write(r.content)

           if mime==".jpg" or mime==".png" or mime==".jpeg" or mime==".bmp" or mime==".webp":
               text=""
               img = cv2.imread("decoded"+mime)
               result = decode(img)
               for i in result:
                   text = (i.data.decode("utf-8"))

               if text!="":
                 print(text)
                 payloadtext = {"text": text,"disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
                 response = requests.post(texturl, json=payloadtext, headers=headers)
                 print(response.text)

                 payloadtext = {"text": advertise, "parse_mode": "html", "disable_web_page_preview": False,
                                "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
                 response = requests.post(texturl, json=payloadtext, headers=headers)
                 break

               elif text == "":
                   payloadtext = {"text": "Sent image does not contain a QR Code. Please send a QR code imgae only ", "parse_mode": "html",
                                  "disable_web_page_preview": False,
                                  "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
                   response = requests.post(texturl, json=payloadtext, headers=headers)
                   print(response.text)

                   payloadtext = {"text": advertise, "parse_mode": "html", "disable_web_page_preview": False,
                                  "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
                   response = requests.post(texturl, json=payloadtext, headers=headers)
                   break
               print(len(a))
           else:
             message="sent file is not a image. It is "+mime+" file"
             payloadtext = {"text": message, "parse_mode": "html",
                            "disable_web_page_preview": False,
                            "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
             response = requests.post(texturl, json=payloadtext, headers=headers)
             print(response.text)
             break

       if messagetext[0:6] == "/start":
           payloadtext = {"text": "Hello "+u1 +"\n\nWelcome to QR Coder/Decoder Bot. Send any text to create QR code or QR code into Text", "parse_mode": "html",
                          "disable_web_page_preview": True,
                          "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
           response = requests.post(texturl, json=payloadtext, headers=headers)
           break

       else:

          if messagetext=="sendad":
              ab = 0
              roww = 2
              okcount=0
              while ab == 0:

                cell_obj1 = sheet_obj.cell(row=roww, column=1)
                cell_obj3 = sheet_obj.cell(row=roww, column=3)
                cell_obj4 = sheet_obj.cell(row=roww, column=4)


                if  len(str(cell_obj1.value)) != 0 and cell_obj4.value!="y":
                  payloadtext = {"text": advertise1, "parse_mode": "html", "disable_web_page_preview": False,
                                 "disable_notification": False, "reply_to_message_id": None, "chat_id": cell_obj1.value}
                  response = requests.post(texturl, json=payloadtext, headers=headers)
                  #print(response.text)
                  capt=response.text
                  status=capt[0:40].find("true")
                  if status!=-1:
                      cell_obj4.value="y"
                      okcount=okcount+1
                      print(str(cell_obj1.value) +"   "+str(okcount))
                      time.sleep(0.5)
                      wb_obj.save(user)

                if cell_obj1.value == "" or cell_obj1.value == None :
                    #wb_obj.close(user)
                    payloadtext = {"text": "Ad sent successfully to : "+str(okcount)+ " persons", "parse_mode": "html", "disable_web_page_preview": False,
                                   "disable_notification": False, "reply_to_message_id": None,
                                   "chat_id": numbertext}
                    response = requests.post(texturl, json=payloadtext, headers=headers)
                    break
                roww = roww + 1

          spcfind=messagetext.find("entities")
          if spcfind!=-1:
             messagetext=messagetext[0:spcfind-4]
          if len(messagetext)>2800:
              message="Input Text must me less that 2800 charcters. Your text is of "+str(len(messagetext))+" characters"
              payloadtext = {"text": len(messagetext), "parse_mode": "html",
                             "disable_web_page_preview": False,
                             "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
              response = requests.post(texturl, json=payloadtext, headers=headers)
              print(response.text)

              payloadtext = {"text": advertise, "parse_mode": "html", "disable_web_page_preview": False,
                             "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
              response = requests.post(texturl, json=payloadtext, headers=headers)
              break
          img = qrcode.make(messagetext)
          img.save("qrcode.jpg")

          file ="qrcode.jpg"

          files = {'photo': open(file, 'rb')}
          response = requests.post(photourl + "?chat_id={}".format(numbertext), files=files)
          print(response.text)

          payloadtext = {"text": advertise, "parse_mode": "html", "disable_web_page_preview": False,
                         "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
          response = requests.post(texturl, json=payloadtext, headers=headers)
          break
       tempstr = tempstr[endpos1+4 :lenc]


    if int(updatetext1)>=int(updatetext):
      updatefile = open("updateid.txt", 'w')
      updatetext = int(updatetext1)+1
      updatefile.write(str(updatetext))
      updatefile.close()

    healthtime1 = int(time.time())
    # print(healthtime1)
    if healthtime1 - healthtime > 1800:
        payloadtext = {"text": healthmessage, "parse_mode": "html", "disable_web_page_preview": False,
                       "disable_notification": False, "reply_to_message_id": None, "chat_id": replyto}
        response = requests.post(texturl, json=payloadtext, headers=headers)
        healthtime = healthtime1

    print("Offset ID in Text File updated succesfully")
